"""
Kotak Neo Dashboard - Web app showing holdings, positions, orders, trades, limits.

Run:
    pip install flask pyotp python-dotenv
    python app.py
Open: http://localhost:5000
"""
import os
import json
import re
import threading
import time
import traceback
from datetime import datetime, timedelta
from flask import Flask, render_template, jsonify, redirect, url_for, request, Response

from backend.utils import IST, now_ist
from backend.kotak.client import (
    _state, login, ensure_client, safe_call,
    append_history, read_history, HISTORY_FILE,
)
from backend.kotak.instruments import (
    SCRIPS, find_scrip,
    INDEX_OPTIONS_CONFIG, _option_universe,
    _fetch_index_fo_universe, _parse_item_strike, _parse_item_expiry_date,
)
from backend.quotes import (
    fetch_quotes, fetch_option_quotes, fetch_future_quotes, build_option_chain,
    build_all_option_tokens, _feed,
)
from backend.snapshot import _store as _snapshot
from backend.storage.trades import (
    LEDGER_FILE, read_trade_ledger, write_trade_ledger, next_trade_id,
)
from backend.storage.orders import ORDERS_FILE, append_order, read_orders
from backend.storage.blocked import (
    append_blocked, read_recent_blocked, read_blocked_since,
    read_blocked_page,
)
from backend import config_loader
from backend.strategy.gann import (
    GANN_STEP, SELL_LEVELS, BUY_LEVELS, LEVEL_COLORS,
    BUY_LEVEL_ORDER, SELL_LEVEL_ORDER,
    gann_levels, nearest_gann_level, compute_target_level_reached,
)
from backend.strategy.common import update_open_trades_mfe, _auto_in_hours
from backend.strategy.options import (
    AUTO_OPTION_STRATEGY_ENABLED, LIVE_MODE,
    _option_auto_state, option_auto_strategy_tick,
)
from backend.strategy.futures import (
    AUTO_FUTURE_STRATEGY_ENABLED,
    _future_auto_state, future_auto_strategy_tick,
)
from backend.strategy.paper_book import (
    paper_options_tick, paper_futures_tick,
)
from backend.safety.kill_switch import (
    is_halted, halt, halt_info,
    is_engine_halted, halt_engine, engine_halt_info,
)
from backend.safety.orders import (
    place_order_safe,
    RESULT_OK, RESULT_PAPER, RESULT_BLOCKED_HALTED,
    RESULT_BLOCKED_MARGIN, RESULT_KOTAK_ERROR,
)
from backend.safety.audit import audit, read_audit_tail, read_audit_page

app = Flask(__name__,
            template_folder="frontend/templates",
            static_folder="frontend/static")


# Jinja filter — render numbers as Indian rupees with thousand separators.
# Used by tile P&L, broker columns, anywhere a numeric ₹ value is shown.
# `value | inr`           -> "₹12,345.68"
# `value | inr(signed=True)` -> "+₹12,345.68" / "-₹1,234.56"
@app.template_filter("inr")
def _jinja_inr(value, signed=False):
    try:
        n = float(value)
    except (TypeError, ValueError):
        return "—"
    if signed:
        sign = "+" if n >= 0 else "-"
        return f"{sign}\u20B9{abs(n):,.2f}"
    return f"\u20B9{n:,.2f}"


def compute_stats(trades):
    open_n   = sum(1 for t in trades if t.get("status") == "OPEN")
    closed_n = sum(1 for t in trades if t.get("status") == "CLOSED")
    wins   = sum(1 for t in trades if t.get("status") == "CLOSED"
                 and (t.get("pnl_points") or 0) > 0)
    losses = sum(1 for t in trades if t.get("status") == "CLOSED"
                 and (t.get("pnl_points") or 0) < 0)
    total_pnl = 0.0
    for t in trades:
        if t.get("status") == "CLOSED" and t.get("pnl_points") is not None:
            try:
                total_pnl += float(t["pnl_points"]) * int(t.get("qty", 1))
            except (TypeError, ValueError):
                pass
    return {
        "total": len(trades), "open": open_n, "closed": closed_n,
        "wins": wins, "losses": losses,
        "total_pnl_points": round(total_pnl, 2),
        # Legacy keys still referenced by older code paths:
        "active": open_n, "pnl": round(total_pnl, 2),
    }


# Date-range filter for the ledger pages (paper + real). Both ledgers store
# `date` as 'YYYY-MM-DD'. The filter is a closed interval on that field,
# computed in IST (now_ist().date()) so weekends / late-night browsing still
# show the trader's intuitive "today". Trades-with-no-date are dropped.
#
# Used by /paper-trades, /trades, /paper-trades.xlsx, /trades.xlsx so the
# table view, the stats tile, and the Excel export all see the same slice.
def _filter_trades_by_range(trades, range_key, custom_date):
    """Return rows whose `date` falls in the requested range.

    range_key: 'today' | 'yesterday' | 'week' | 'month' | 'all' | 'custom'
    custom_date: 'YYYY-MM-DD' string (only honoured when range_key=='custom')
    """
    if not range_key or range_key == "all":
        return list(trades)
    today = now_ist().date()
    if range_key == "today":
        start = end = today
    elif range_key == "yesterday":
        start = end = today - timedelta(days=1)
    elif range_key == "week":
        # Monday-of-this-week through today, inclusive. weekday(): Mon=0..Sun=6.
        start = today - timedelta(days=today.weekday())
        end = today
    elif range_key == "month":
        start = today.replace(day=1)
        end = today
    elif range_key == "custom":
        try:
            start = end = datetime.strptime(
                (custom_date or "")[:10], "%Y-%m-%d").date()
        except (TypeError, ValueError):
            # Bad/missing date — fall back to unfiltered so the user can still
            # see something rather than an empty page.
            return list(trades)
    else:
        return list(trades)
    out = []
    for t in trades:
        d = t.get("date") or ""
        if not d:
            continue
        try:
            td = datetime.strptime(d[:10], "%Y-%m-%d").date()
        except (TypeError, ValueError):
            continue
        if start <= td <= end:
            out.append(t)
    return out


def _resolve_date_range(args):
    """Pull (range_key, custom_date) from a request.args.
    Default = 'today'. If a `date=` param is present without an explicit
    `range=`, treat it as a custom-date filter (so the audit/blockers
    legacy single-date links keep working unchanged)."""
    range_key   = (args.get("range") or "").strip().lower()
    custom_date = (args.get("date")  or "").strip()
    if not range_key:
        range_key = "custom" if custom_date else "today"
    return range_key, custom_date


def fmt_duration(seconds):
    if seconds is None or seconds < 0:
        return ""
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


# Tabs are grouped left-to-right by purpose: Live (what's trading right now),
# Books (ledgers), Risk (safety/forensics), Account (broker passthrough +
# config). Order matches Ganesh's daily workflow — live screens first, then
# books to review, then risk + account at the end.
TABS = [
    # Live — what the bot is trading right now
    {"key": "gann",         "url": "/gann",         "label": "Gann Trader"},
    {"key": "options",      "url": "/options",      "label": "Options"},
    {"key": "futures",      "url": "/futures",      "label": "Futures"},
    # Books — ledgers (paper next to real, real first)
    {"key": "trades",       "url": "/trades",       "label": "Trade Log"},
    {"key": "paper_trades", "url": "/paper-trades", "label": "Paper Log"},
    # Risk — safety / forensics
    {"key": "blockers",     "url": "/blockers",     "label": "Blockers"},
    {"key": "audit",        "url": "/audit",        "label": "Audit"},
    # Account — broker passthrough + config + login history
    {"key": "holdings",     "url": "/",             "label": "Holdings"},
    {"key": "positions",    "url": "/positions",    "label": "Positions"},
    {"key": "config",       "url": "/config",       "label": "Config"},
    {"key": "history",      "url": "/history",      "label": "Login History"},
]






# C.1 — curated column maps for broker passthrough pages. The Kotak SDK
# returns rows with mixed casing (trdSym vs tradingSymbol vs tSym, prc vs
# price, etc.) and dozens of fields per row, most of them noise. The
# auto-column path (used previously) dumped 30+ columns and made the table
# unreadable on a laptop.
#
# Each entry is (Label, [field-name aliases]) — render() picks the first
# alias that has a non-empty value in each row. Aliases compiled from
# Kotak's REST docs and existing code in backend/safety/positions.py +
# backend/safety/orders.py. If the SDK adds or renames a field, only this
# table needs an update.
BROKER_COLUMN_MAPS = {
    "holdings": [
        # Bug 3 — alias list expanded to cover the additional field names
        # the Kotak Neo SDK has been seen to return for holdings rows.
        # Aliases are evaluated first-non-empty, so adding fallbacks is
        # always safe — they only matter when earlier ones are missing.
        ("Symbol",   ["displaySymbol", "tradingSymbol", "trdSym",
                      "tSym", "symbol", "instrumentName"]),
        ("Exchange", ["exchangeSegment", "exSeg", "exchange", "exch"]),
        ("Qty",      ["quantity", "qty", "totalQty", "displayQty",
                      "holdingsQty", "holdQty", "sellableQty"]),
        ("Avg Price",["averagePrice", "avgPrice", "avgPrc", "buyAvg",
                      "costAvg"]),
        ("LTP",      ["closingPrice", "lastPrice", "ltp", "lstPrc",
                      "mktPrice"]),
        ("P&L",      ["unRealizedPnL", "pnl", "dayChange", "mtm",
                      "dayPnL", "pnL"]),
    ],
    "positions": [
        ("Symbol",   ["trdSym", "tradingSymbol", "tSym", "symbol"]),
        ("Product",  ["product", "prdt", "prod"]),
        ("Net Qty",  ["netQty", "netTrdQty", "netTrdQtyLot",
                      "flBuyQty"]),
        ("Avg Price",["averagePrice", "avgPrice", "avgPrc", "netAvg"]),
        ("LTP",      ["lastPrice", "ltp", "lstPrc"]),
        ("MTM P&L",  ["mtmPnl", "pnl", "rpnl", "urpnl"]),
    ],
    "orders": [
        ("Time",     ["orderTime", "ordEntTm", "exchTime", "ordEntryTm"]),
        ("Symbol",   ["trdSym", "tradingSymbol", "tSym", "symbol"]),
        ("Side",     ["transactionType", "trnsTp", "buyOrSell"]),
        ("Qty",      ["quantity", "qty", "ordQty"]),
        ("Price",    ["price", "prc", "limitPrice"]),
        ("Status",   ["status", "ordSt", "orderStatus"]),
        ("Order ID", ["orderId", "nOrdNo", "exchOrderId"]),
    ],
    "trade-book": [
        ("Time",     ["tradeTime", "trdTm", "exchTime", "fillTime"]),
        ("Symbol",   ["trdSym", "tradingSymbol", "tSym", "symbol"]),
        ("Side",     ["transactionType", "trnsTp", "buyOrSell"]),
        ("Qty",      ["filledQty", "trdQty", "qty", "fillQty"]),
        ("Price",    ["tradePrice", "trdPrc", "fillPrice", "avgPrc"]),
        ("Order ID", ["orderId", "nOrdNo", "exchOrderId"]),
    ],
    # /limits is a single flat dict, not a list. We still keep a curated
    # field list and let render() flatten it into a one-row table. Field
    # names verified against a live Kotak Neo response: keys are mixed
    # camel/Pascal case with `Prsnt` suffix (American spelling — Kotak's
    # API uses "Realized" / "Unrealized", not the British forms).
    "limits": [
        ("Notional Cash",    ["NotionalCash", "AuxNotionalCash"]),
        ("Collateral",       ["Collateral", "CollateralValue", "RmsCollateral"]),
        ("Margin Used",      ["MarginUsedPrsnt", "MarginUsed", "AmountUtilizedPrsnt"]),
        ("Net",              ["Net", "netAvailable"]),
        ("Realised P&L",     ["RealizedMtomPrsnt", "CashRlsMtomPrsnt", "FoRlsMtomPrsnt"]),
        ("Unrealised P&L",   ["UnrealizedMtomPrsnt", "CashUnRlsMtomPrsnt", "FoUnRlsMtomPrsnt"]),
    ],
}


# Kotak's REST endpoints return an error envelope when the request fails
# (auth not yet ready, market closed, scrip mismatch, etc.) instead of
# the expected list/dict shape. The envelope shape is {stat, stCode, desc,
# errMsg}. The render() path used to project this through the column map
# and produce an empty row — visually indistinguishable from "no data" —
# so the user couldn't tell whether the broker rejected the call. Treat
# any row whose key set is a subset of this envelope as an error and
# surface errMsg/desc as the view-error banner instead.
_KOTAK_ERROR_ENVELOPE_KEYS = {"stat", "stCode", "desc", "errMsg"}


def _is_kotak_error_envelope(row):
    if not isinstance(row, dict) or not row:
        return False
    keys = set(row.keys())
    # All keys must come from the envelope set; allow partials like
    # {stat, errMsg} too. A real data row always has at least one key
    # outside this set.
    return keys.issubset(_KOTAK_ERROR_ENVELOPE_KEYS)


def _pick(row, aliases):
    """Return the first non-empty value among `aliases` in `row`. Empty
    strings and None count as missing; 0 / 0.0 / False are kept (so a zero
    qty or a false flag still renders)."""
    if not isinstance(row, dict):
        return ""
    for k in aliases:
        if k in row:
            v = row[k]
            if v is None or v == "":
                continue
            return v
    return ""


# C.2 — extract the underlying ticker from a trading symbol so we can group
# all legs of the same name on /positions. Kotak symbols look like
# `NIFTY01MAY2625500CE`, `BANKNIFTY24APR26FUT`, `RELIANCE-EQ` — the
# underlying is always the leading alpha run (everything before the first
# digit or hyphen). Returns "" for blanks; the caller's sort puts those last.
_UNDERLYING_RE = re.compile(r"^([A-Z]+)")


def _extract_underlying(symbol):
    if not symbol:
        return ""
    m = _UNDERLYING_RE.match(str(symbol).upper())
    return m.group(1) if m else ""


def render(active, heading, data, error):
    rows = []
    cols = []
    col_map = BROKER_COLUMN_MAPS.get(active)
    if data:
        # Normalise: SDK returns list of dicts (each row) or a dict-of-scalars
        if isinstance(data, list):
            rows = data
        elif isinstance(data, dict):
            # Some endpoints (limits) return a single flat dict - show as one row
            rows = [data]

        # C.1 — promote a Kotak error envelope to view_error so the user
        # sees "scrip not found" / "session expired" instead of a blank
        # table. Single-row case only; if a real list happens to contain
        # one error row we still want to skip it (treat as no data).
        if rows and len(rows) == 1 and _is_kotak_error_envelope(rows[0]):
            env = rows[0]
            error = (env.get("errMsg") or env.get("desc")
                     or f"Kotak API error (stCode={env.get('stCode')})")
            rows = []

        if col_map:
            # C.1 — curated path. Project each broker row into a dict with
            # exactly the labels in the column map, looking up via aliases.
            # This collapses 30+ raw fields into ~6 readable columns.
            cols = [label for (label, _aliases) in col_map]
            projected = []
            for r in rows:
                projected.append({label: _pick(r, aliases)
                                  for (label, aliases) in col_map})
            rows = projected
            # C.2 — group positions by underlying so all NIFTY legs (and all
            # BANKNIFTY legs, etc.) cluster together. Kite shows a combined
            # group header; we settle for an explicit column + sort, which
            # gives the same scannability without rebuilding the table layout.
            if active == "positions":
                for r in rows:
                    r["Underlying"] = _extract_underlying(r.get("Symbol", ""))
                rows.sort(key=lambda r: (r.get("Underlying") or "~",
                                         str(r.get("Symbol", ""))))
                cols = ["Underlying"] + cols
        elif rows and isinstance(rows[0], dict):
            # Fallback (no curated map for this page): legacy auto-column
            # behaviour. Collect all keys across rows, preserve first-row
            # order. Only used for unknown `active` values.
            seen = []
            for r in rows:
                for k in r.keys():
                    if k not in seen:
                        seen.append(k)
            cols = seen

    return render_template(
        "base.html",
        tabs=TABS,
        active=active,
        heading=heading,
        rows=rows,
        cols=cols,
        view_error=error,
        greeting=_state.get("greeting"),
        login_time=_state["login_time"].strftime("%H:%M:%S IST") if _state.get("login_time") else None,
        error=_state.get("error"),
    )


# D.1 — broker passthrough cache. Each Kotak REST round-trip is 200–700 ms
# p95 (see docs/PERF_REPORT.md), and these pages are commonly tab-rotated
# every few seconds. A small per-method 5 s TTL collapses the burst-refresh
# pattern to one REST hit every 5 s while the auto-strategy still sees fresh
# data on its own tick (it doesn't read these endpoints).
_BROKER_CACHE = {}      # name -> (ts, data, err)
_BROKER_CACHE_LOCK = threading.Lock()
_BROKER_CACHE_TTL = 5.0


def _cached_safe_call(name, fn, *args, **kwargs):
    """safe_call with a per-name 5 s TTL cache. Cache key includes kwargs
    so different argument sets don't collide (e.g. /limits could ask for
    different product slices). Errors are cached too — that's intentional;
    a downstream 5 s burst of identical errors avoids hammering Kotak when
    the breaker has just opened."""
    key = (name, tuple(sorted(kwargs.items())), args)
    now = time.time()
    with _BROKER_CACHE_LOCK:
        ent = _BROKER_CACHE.get(key)
        if ent and (now - ent[0]) < _BROKER_CACHE_TTL:
            return ent[1], ent[2]
    data, err = safe_call(fn, *args, **kwargs)
    with _BROKER_CACHE_LOCK:
        _BROKER_CACHE[key] = (now, data, err)
    return data, err


@app.route("/")
def holdings_view():
    try:
        client = ensure_client()
        data, err = _cached_safe_call("holdings", client.holdings)
        return render("holdings", "Portfolio Holdings", data, err)
    except Exception as e:
        return render("holdings", "Portfolio Holdings", None, traceback.format_exc())


@app.route("/positions")
def positions_view():
    try:
        client = ensure_client()
        data, err = _cached_safe_call("positions", client.positions)
        return render("positions", "Open Positions", data, err)
    except Exception as e:
        return render("positions", "Open Positions", None, traceback.format_exc())


@app.route("/orders")
def orders_view():
    try:
        client = ensure_client()
        data, err = _cached_safe_call("orders", client.order_report)
        return render("orders", "Order Book", data, err)
    except Exception as e:
        return render("orders", "Order Book", None, traceback.format_exc())


@app.route("/trade-book")
def trade_book_view():
    """Kotak's broker-side executed trade list. Distinct from our internal
    trade ledger (/trades) which records every signal we acted on."""
    try:
        client = ensure_client()
        data, err = _cached_safe_call("trade_book", client.trade_report)
        return render("trade-book", "Trade Book", data, err)
    except Exception as e:
        return render("trade-book", "Trade Book", None, traceback.format_exc())


@app.route("/limits")
def limits_view():
    try:
        client = ensure_client()
        data, err = _cached_safe_call(
            "limits", client.limits,
            segment="ALL", exchange="ALL", product="ALL")
        return render("limits", "Funds & Limits", data, err)
    except Exception as e:
        return render("limits", "Funds & Limits", None, traceback.format_exc())


@app.route("/history")
def history_view():
    return render_template(
        "history.html",
        tabs=TABS,
        active="history",
        history=read_history(),
    )


@app.route("/gann")
def gann_view():
    return render_template(
        "gann.html",
        tabs=TABS,
        active="gann",
        scrips=SCRIPS,
        ucc=os.getenv("KOTAK_UCC", ""),
        level_colors=LEVEL_COLORS,
    )


@app.route("/api/feed-status")
def feed_status_api():
    """WebSocket QuoteFeed diagnostics."""
    # _feed_started and WS_FRESH_SECONDS now live in backend.quotes after
    # the quote pipeline was extracted from app.py. Import lazily so we
    # don't pollute the top-level import block.
    from backend.quotes import _feed_started, WS_FRESH_SECONDS
    return jsonify({
        "started": _feed_started["flag"],
        **_feed.status(),
        "fresh_threshold_seconds": WS_FRESH_SECONDS,
        "ts": now_ist().strftime("%H:%M:%S IST"),
    })


@app.route("/api/health")
def health_api():
    """Strong-API stats: per-method call counts, errors, retries, breaker state."""
    from backend.kotak.api import stats as kotak_stats
    from backend.quotes import _feed_started  # moved out of app.py
    return jsonify({
        "ok": True,
        "kotak": kotak_stats(),
        "feed": _feed.status(),
        "feed_started": _feed_started["flag"],
        "logged_in": _state.get("client") is not None,
        "ts": now_ist().strftime("%H:%M:%S IST"),
    })


@app.route("/api/gann-prices")
def gann_prices_api():
    """O(1) read from SnapshotStore. The producer thread refreshes the
    payload every 2s by calling the same fetch_quotes() pipeline; this
    handler does no I/O of its own."""
    blob, built_at, build_ms = _snapshot.gann_payload()
    age_ms = (time.time() - built_at) * 1000.0 if built_at else -1.0
    resp = Response(blob, mimetype="application/json")
    resp.headers["X-Snapshot-Age-Ms"] = f"{age_ms:.0f}"
    resp.headers["X-Snapshot-Build-Ms"] = f"{build_ms:.0f}"
    return resp


@app.route("/api/gann-live")
def gann_live_api():
    """Sub-second LTP for the Gann page. Reads directly from the WS
    QuoteFeed cache — no REST, no SnapshotStore. Designed to be polled
    every ~500ms so cells update at the same cadence as Kotak's app.

    The slower /api/gann-prices keeps providing levels + stats. This
    endpoint only sends `ltp` + recomputed `nearest_level` so the JS
    can repaint LTP cells without touching the rest of the row.
    """
    from backend.quotes import _quote_cache
    cached = _quote_cache.get("data") or {}
    out = []
    now = time.time()
    for s in SCRIPS:
        sym = s["symbol"]
        tick = _feed.get(s["exchange"], s["token"]) or {}
        ltp = tick.get("ltp")
        ws_age = round(now - tick.get("ts", 0), 2) if tick.get("ts") else None
        # Fall back to last cached LTP if WS hasn't ticked yet (e.g. just-
        # subscribed scrip or off-hours) — better than blanking the cell.
        if ltp is None:
            cached_row = cached.get(sym) or {}
            ltp = cached_row.get("ltp")
        # Use cached levels (rebuilt on REST refresh) to compute nearest.
        cached_row = cached.get(sym) or {}
        levels = cached_row.get("levels") or {"sell": {}, "buy": {}}
        nl, _ = nearest_gann_level({"ltp": ltp, "levels": levels})
        out.append({
            "symbol": sym,
            "ltp": ltp,
            "nearest_level": nl,
            "ws_age": ws_age,
        })
    return jsonify({
        "scrips": out,
        "ts": now_ist().strftime("%H:%M:%S IST"),
    })


# ---------- Options routes ----------
@app.route("/options")
def options_view():
    """Server-side render uses the SnapshotStore so the page comes down with
    the chain already populated and reload/tab-switch is instant. If the
    snapshot hasn't been built yet (first request after boot, before the
    F&O universe is warm) the producer's empty placeholder bubbles up as
    `loading: true` and the template falls back to "Loading…" + JS poller."""
    try:
        blob, _, _ = _snapshot.options_payload()
        initial_data = json.loads(blob)
        if initial_data.get("loading"):
            initial_data = None
    except Exception as e:
        print(f"[options_view] snapshot read failed: {type(e).__name__}: {e}")
        initial_data = None
    return render_template(
        "options.html",
        tabs=TABS,
        active="options",
        indices=[{"key": k, "label": v["label"]} for k, v in INDEX_OPTIONS_CONFIG.items()],
        ucc=os.getenv("KOTAK_UCC", ""),
        initial_data=initial_data,
    )


@app.route("/api/option-prices")
def option_prices_api():
    """O(1) read from SnapshotStore. The producer thread does the heavy
    REST work every 2s; this handler reads the pre-built bytes.

    The auto-strategy tick is owned by the autonomous _strategy_ticker_loop
    daemon — it does NOT run from this request anymore, so refreshing the
    page no longer drives strategy decisions."""
    # On first hit of the day the F&O universe may not yet be warm. Kick
    # off the preload in background so the snapshot fills in on the next
    # producer iteration. Idempotent.
    if not _option_universe.get("loading") and not all(
            i in _option_universe.get("by_index", {})
            for i in INDEX_OPTIONS_CONFIG):
        _option_universe["loading"] = True
        def _warm():
            try:
                _preload_option_universe()
            finally:
                _option_universe["loading"] = False
        threading.Thread(target=_warm, daemon=True).start()
    blob, built_at, build_ms = _snapshot.options_payload()
    age_ms = (time.time() - built_at) * 1000.0 if built_at else -1.0
    resp = Response(blob, mimetype="application/json")
    resp.headers["X-Snapshot-Age-Ms"] = f"{age_ms:.0f}"
    resp.headers["X-Snapshot-Build-Ms"] = f"{build_ms:.0f}"
    return resp


@app.route("/api/snapshot-stats")
def snapshot_stats_api():
    """Diagnostic: how the SnapshotStore producer is keeping up."""
    return jsonify(_snapshot.stats())


# ---------- Futures routes ----------
@app.route("/futures")
def futures_view():
    """Futures dashboard — live LTP, signal, limit-price preview per index,
    plus today's auto futures trades. Read-only; entries fire from the
    autonomous strategy ticker (every 3s)."""
    return render_template(
        "futures.html",
        tabs=TABS,
        active="futures",
        indices=[{"key": k, "label": v["label"]}
                 for k, v in INDEX_OPTIONS_CONFIG.items()],
        ucc=os.getenv("KOTAK_UCC", ""),
    )


@app.route("/api/future-prices")
def future_prices_api():
    """O(1) read from SnapshotStore. Same hot-cache pattern as the gann +
    options endpoints — the producer thread maintains the futures payload
    on a 2s tick."""
    blob, built_at, build_ms = _snapshot.futures_payload()
    age_ms = (time.time() - built_at) * 1000.0 if built_at else -1.0
    resp = Response(blob, mimetype="application/json")
    resp.headers["X-Snapshot-Age-Ms"] = f"{age_ms:.0f}"
    resp.headers["X-Snapshot-Build-Ms"] = f"{build_ms:.0f}"
    return resp


# ---------- Trade ledger routes ----------
@app.route("/api/trades")
def trades_api():
    trades = read_trade_ledger()
    return jsonify({"trades": trades, "stats": compute_stats(trades)})


@app.route("/trades")
def trades_view():
    """Trade Log page: every signal acted on, with a 'Download as Excel' button.
    Includes both LIVE rows (real Kotak orders) and any legacy paper rows.

    Date strip (Today/Yesterday/Week/Month/All/Custom) is applied server-side.
    Default = Today. Stats tile + table + Excel link all reflect the same
    filtered slice. Independent of /paper-trades — different ledger file."""
    trades = read_trade_ledger()
    range_key, custom_date = _resolve_date_range(request.args)
    trades_filtered = _filter_trades_by_range(trades, range_key, custom_date)
    trades_sorted = sorted(
        trades_filtered,
        key=lambda t: (t.get("date") or "", t.get("entry_time") or ""),
        reverse=True,
    )
    for t in trades_sorted:
        t["duration_str"] = fmt_duration(t.get("duration_seconds"))
    return render_template(
        "trade_ledger.html",
        tabs=TABS,
        active="trades",
        trades=trades_sorted,
        stats=compute_stats(trades_filtered),
        cfg=config_loader.get(),
        range_key=range_key,
        custom_date=custom_date,
    )


@app.route("/api/paper-trades-live")
def paper_trades_live_api():
    """Live LTP + recomputed P&L for every OPEN paper row.
    Polled by paper_trades.html so Ganesh can watch the trail SL move
    in real-time vs the spot/option LTP.

    Read order:
      1. WS feed via stored instrument_token+exchange (cheapest, freshest)
      2. option/future REST quote caches (fallback for legacy rows)
    The WS path is necessary for option strikes that drift out of the
    ATM window — the REST cache only covers strikes near current ATM,
    so an OPEN trade at an old strike would otherwise show null.
    """
    from backend.storage.paper_ledger import read_paper_ledger
    from backend.quotes import (
        _option_quote_cache, _future_quote_cache, _quote_cache,
    )
    opt_data = _option_quote_cache.get("data") or {}
    fut_data = _future_quote_cache.get("data") or {}
    spot_cache = _quote_cache.get("data") or {}

    # Build {underlying_name -> live_spot} once per request. Ganesh
    # wants to see spot move alongside the option LTP so he can watch
    # the variant-D trail SL trigger in real time.
    spot_by_underlying = {}
    sym_to_underlying = {
        cfg["spot_symbol_key"]: idx_name
        for idx_name, cfg in INDEX_OPTIONS_CONFIG.items()
    }
    for s in SCRIPS:
        if s["symbol"] not in sym_to_underlying:
            continue
        tick = _feed.get(s["exchange"], s["token"]) or {}
        spot = tick.get("ltp")
        if spot is None:
            spot = (spot_cache.get(s["symbol"]) or {}).get("ltp")
        spot_by_underlying[sym_to_underlying[s["symbol"]]] = spot

    rows = read_paper_ledger()
    out = []
    for t in rows:
        if t.get("status") != "OPEN":
            continue
        ltp = None
        # Path 1: WS direct read using stored token (fresh sub-second).
        token = t.get("instrument_token")
        exch  = t.get("exchange_segment")
        if token and exch:
            tick = _feed.get(exch, str(token)) or {}
            ltp = tick.get("ltp")
        # Path 2: fallback to REST cache for legacy rows w/o token.
        if ltp is None:
            if t.get("asset_type") == "option":
                q = opt_data.get(t.get("option_key")) or {}
                ltp = q.get("ltp")
            elif t.get("asset_type") == "future":
                q = fut_data.get(t.get("underlying")) or {}
                ltp = q.get("ltp")
        pnl_pts = pnl_pct = None
        entry_price = t.get("entry_price")
        if ltp is not None and entry_price:
            if t.get("order_type") == "BUY":
                pnl_pts = round(float(ltp) - float(entry_price), 2)
            else:
                pnl_pts = round(float(entry_price) - float(ltp), 2)
            pnl_pct = round((pnl_pts / float(entry_price)) * 100, 2)
        out.append({
            "id": t.get("id"),
            "ltp": ltp,
            "spot": spot_by_underlying.get(t.get("underlying")),
            "pnl_points": pnl_pts,
            "pnl_pct": pnl_pct,
            "trail_sl_price": t.get("trail_sl_price"),
            "trail_high_rung": t.get("trail_high_rung"),
        })
    return jsonify({"trades": out, "ts": now_ist().strftime("%H:%M:%S IST")})


@app.route("/paper-trades")
def paper_trades_view():
    """Paper Book page — independent ledger of virtual paper trades.

    Date strip (Today/Yesterday/Week/Month/All/Custom) is applied
    server-side. Default = Today. Stats tile + table + Excel link all
    reflect the same filtered slice."""
    from backend.storage.paper_ledger import read_paper_ledger
    rows = read_paper_ledger()
    range_key, custom_date = _resolve_date_range(request.args)
    rows_filtered = _filter_trades_by_range(rows, range_key, custom_date)
    rows_sorted = sorted(
        rows_filtered,
        key=lambda t: (t.get("date") or "", t.get("entry_time") or ""),
        reverse=True,
    )
    for t in rows_sorted:
        t["duration_str"] = fmt_duration(t.get("duration_seconds"))
    return render_template(
        "paper_trades.html",
        tabs=TABS,
        active="paper_trades",
        trades=rows_sorted,
        stats=compute_stats(rows_filtered),
        cfg=config_loader.get(),
        range_key=range_key,
        custom_date=custom_date,
    )


@app.route("/paper-trades.xlsx")
def paper_trades_xlsx():
    """Export the paper ledger as a formatted Excel file."""
    from backend.storage.paper_ledger import read_paper_ledger
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Paper Ledger"
    # Phase 3: Engine column (current/reverse) inserted after Order Type so
    # Ganesh can filter the export by engine in Excel. Legacy rows that
    # predate Phase 2 don't have an "engine" key and get rendered as
    # 'current' to match the HTML table behaviour.
    headers = ["Date", "Scrip", "Order Type", "Engine", "Entry Time (IST)",
               "Entry Price", "Target Level Reached", "Max/Min Target Price",
               "Exit Time (IST)", "Exit Price", "Exit Reason",
               "P&L Points", "P&L %", "Duration"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="FFEB3B")
    for c, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    sell_fill = PatternFill("solid", fgColor="F8CBAD")
    buy_fill  = PatternFill("solid", fgColor="C6EFCE")
    pos_font  = Font(color="006100")
    neg_font  = Font(color="9C0006")
    # Honour the same date-range filter as the /paper-trades page so the
    # Excel download contains exactly the rows the user is currently viewing.
    range_key, custom_date = _resolve_date_range(request.args)
    _paper_rows = _filter_trades_by_range(read_paper_ledger(),
                                          range_key, custom_date)
    for t in _paper_rows:
        ws.append([
            t.get("date", ""),
            t.get("scrip", ""),
            t.get("order_type", ""),
            t.get("engine") or "current",
            t.get("entry_time", ""),
            t.get("entry_price", ""),
            t.get("target_level_reached", "") or "",
            t.get("max_min_target_price", "") or "",
            t.get("exit_time", "") or "",
            t.get("exit_price", "") or "",
            t.get("exit_reason", "") or "",
            t.get("pnl_points", "") if t.get("pnl_points") is not None else "",
            t.get("pnl_pct", "") if t.get("pnl_pct") is not None else "",
            fmt_duration(t.get("duration_seconds")),
        ])
        r = ws.max_row
        otype_cell = ws.cell(row=r, column=3)
        if t.get("order_type") == "SELL":
            otype_cell.fill = sell_fill
        else:
            otype_cell.fill = buy_fill
        otype_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        try:
            pl = float(t.get("pnl_points") or 0)
            # P&L points is now col 12, P&L % col 13 after Engine insert.
            ws.cell(row=r, column=12).font = pos_font if pl >= 0 else neg_font
            ws.cell(row=r, column=13).font = pos_font if pl >= 0 else neg_font
        except (TypeError, ValueError):
            pass
    widths = [12, 12, 12, 10, 18, 14, 20, 20, 18, 12, 14, 12, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    os.makedirs(data_dir, exist_ok=True)
    out = os.path.join(data_dir, "_paper_ledger_export.xlsx")
    wb.save(out)
    with open(out, "rb") as f:
        data = f.read()
    # Filename includes the range so users can tell which slice they downloaded.
    # 'today_20260429.xlsx' / 'week.xlsx' / 'custom_20260415.xlsx' / 'all.xlsx'
    today = now_ist().strftime("%Y%m%d")
    if range_key == "custom" and custom_date:
        suffix = f"custom_{custom_date.replace('-', '')}"
    elif range_key == "today":
        suffix = f"today_{today}"
    elif range_key in ("yesterday", "week", "month", "all"):
        suffix = range_key
    else:
        suffix = today
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename=paper_ledger_{suffix}.xlsx"},
    )


@app.route("/trades.xlsx")
def trades_xlsx():
    """Export the trade ledger as a formatted Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Ledger"
    # Phase 3: Engine column (current/reverse) — see /paper-trades.xlsx for
    # the rationale. Inserted after Order Type; legacy rows fall back to
    # 'current' to match the HTML rendering.
    headers = ["Date", "Scrip", "Order Type", "Engine", "Entry Time (IST)",
               "Entry Price", "Target Level Reached", "Max/Min Target Price",
               "Exit Time (IST)", "Exit Price", "Exit Reason",
               "P&L Points", "P&L %", "Duration"]
    ws.append(headers)
    # Header styling
    hdr_fill = PatternFill("solid", fgColor="FFEB3B")
    for c, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    sell_fill = PatternFill("solid", fgColor="F8CBAD")
    buy_fill  = PatternFill("solid", fgColor="C6EFCE")
    pos_font  = Font(color="006100")
    neg_font  = Font(color="9C0006")
    # Honour the same date-range filter as the /trades page so the Excel
    # download contains exactly the rows the user is currently viewing.
    range_key, custom_date = _resolve_date_range(request.args)
    _real_rows = _filter_trades_by_range(read_trade_ledger(),
                                         range_key, custom_date)
    for t in _real_rows:
        ws.append([
            t.get("date", ""),
            t.get("scrip", ""),
            t.get("order_type", ""),
            t.get("engine") or "current",
            t.get("entry_time", ""),
            t.get("entry_price", ""),
            t.get("target_level_reached", "") or "",
            t.get("max_min_target_price", "") or "",
            t.get("exit_time", "") or "",
            t.get("exit_price", "") or "",
            t.get("exit_reason", "") or "",
            t.get("pnl_points", "") if t.get("pnl_points") is not None else "",
            t.get("pnl_pct", "") if t.get("pnl_pct") is not None else "",
            fmt_duration(t.get("duration_seconds")),
        ])
        r = ws.max_row
        # Colour the Order Type cell
        otype_cell = ws.cell(row=r, column=3)
        if t.get("order_type") == "SELL":
            otype_cell.fill = sell_fill
        else:
            otype_cell.fill = buy_fill
        otype_cell.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        # P&L colouring (cols shifted right by 1 after Engine insert)
        try:
            pl = float(t.get("pnl_points") or 0)
            ws.cell(row=r, column=12).font = pos_font if pl >= 0 else neg_font
            ws.cell(row=r, column=13).font = pos_font if pl >= 0 else neg_font
        except (TypeError, ValueError):
            pass
    # Column widths
    widths = [12, 12, 12, 10, 18, 14, 20, 20, 18, 12, 14, 12, 10, 12]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Write the xlsx to data/ (kept out of source listing) then stream it back.
    data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
    os.makedirs(data_dir, exist_ok=True)
    out = os.path.join(data_dir, "_trade_ledger_export.xlsx")
    wb.save(out)
    with open(out, "rb") as f:
        data = f.read()
    # Filename includes the range so users can tell which slice they downloaded.
    today = now_ist().strftime("%Y%m%d")
    if range_key == "custom" and custom_date:
        suffix = f"custom_{custom_date.replace('-', '')}"
    elif range_key == "today":
        suffix = f"today_{today}"
    elif range_key in ("yesterday", "week", "month", "all"):
        suffix = range_key
    else:
        suffix = today
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename=trade_ledger_{suffix}.xlsx"},
    )


# ---------- Blockers (refused order attempts) ----------
@app.route("/blockers")
def blockers_view():
    """Page showing every order the safety wrapper refused. Paginated
    server-side to keep the response small even when the JSONL store has
    grown to thousands of rows. Query params: page=N, date=YYYY-MM-DD,
    kind=ENTRY|EXIT, source=auto_options|auto_futures|... (F.4 filters)."""
    page = request.args.get("page", 1)
    date = request.args.get("date") or ""
    kind = request.args.get("kind") or ""
    source = request.args.get("source") or ""
    p = read_blocked_page(page=page, page_size=50, date=date,
                          kind=kind, source=source)
    return render_template(
        "blockers.html",
        tabs=TABS,
        active="blockers",
        blocks=p["items"],
        pagination=p,
        date_filter=date,
        kind_filter=kind,
        source_filter=source,
        distinct_kinds=p.get("distinct_kinds", []),
        distinct_sources=p.get("distinct_sources", []),
    )


@app.route("/api/recent-blocks")
def recent_blocks_api():
    """Toaster poll endpoint. Returns blocked-attempt records strictly newer
    than `since` (ISO timestamp). Browsers tail this every few seconds and
    pop a red toast for each new record."""
    since_ts = (request.args.get("since") or "").strip()
    return jsonify({
        "blocks": read_blocked_since(since_ts),
        "ts": now_ist().isoformat(),
    })


@app.route("/api/blocked-list")
def blocked_list_api():
    """Live-table feed for the /blockers page. Returns ONE PAGE of records
    (newest first) plus pagination metadata. The page poller asks for the
    page the user is currently viewing — it never pulls the full file
    again. That's the whole pagination win: small payload + bounded DOM.
    Honours the same kind/source filter params as /blockers so the live
    poll stays consistent with what the user is currently looking at."""
    page = request.args.get("page", 1)
    date = request.args.get("date") or ""
    kind = request.args.get("kind") or ""
    source = request.args.get("source") or ""
    p = read_blocked_page(page=page, page_size=50, date=date,
                          kind=kind, source=source)
    return jsonify({
        "blocks": p["items"],
        "page": p["page"],
        "pages": p["pages"],
        "page_size": p["page_size"],
        "total": p["total"],
        "ts": now_ist().isoformat(),
    })


# ---------- Config (user-tunable strategy params) ----------
@app.route("/config")
def config_view():
    """Render the strategy-config form. All values come from config.yaml
    via config_loader (which hot-reloads on file mtime change)."""
    return render_template(
        "config.html",
        tabs=TABS,
        active="config",
        cfg=config_loader.get(),
    )


@app.route("/api/config", methods=["GET", "POST"])
def config_api():
    """GET returns the current config. POST validates + writes config.yaml.
    The strategy tick picks up changes on its next pass (within ~3s)."""
    if request.method == "GET":
        return jsonify({"ok": True, "config": config_loader.get()})
    payload = request.get_json(force=True, silent=True) or {}
    try:
        saved = config_loader.save(payload)
        audit("CONFIG_SAVED", keys=list(saved.keys()))
        return jsonify({"ok": True, "config": saved})
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    except Exception as e:
        return jsonify({"ok": False,
                        "error": f"{type(e).__name__}: {e}"}), 500


@app.route("/api/margin-summary")
def margin_summary_api():
    """Return available cash so the order ticket can show margin headroom."""
    try:
        client = ensure_client()
        data, err = safe_call(client.limits, segment="ALL", exchange="ALL", product="ALL")
        if err:
            return jsonify({"error": err, "available": None})
        # limits may return dict-of-fields or list-of-rows. Surface the most
        # useful field. Kotak typically uses Net or CashAvailable / Net.
        avail = None
        if isinstance(data, dict):
            for k in ("Net", "net", "CashAvailable", "cashAvailable",
                     "AvailableCash", "availableCash", "DepositValue"):
                if k in data:
                    try:
                        avail = float(data[k])
                        break
                    except (TypeError, ValueError):
                        pass
        return jsonify({"available": avail, "raw": data, "error": None})
    except Exception as e:
        return jsonify({"error": str(e), "available": None})


@app.route("/api/place-order", methods=["POST"])
def place_order_api():
    """Place a real order with Kotak. Records every attempt to orders_log.json."""
    payload = request.get_json(force=True, silent=True) or {}
    symbol      = (payload.get("symbol") or "").strip()
    side        = (payload.get("side") or "").strip().upper()      # B or S
    qty         = str(payload.get("qty") or "").strip()
    order_type  = (payload.get("order_type") or "L").strip().upper()  # L / MKT
    price       = str(payload.get("price") or "0").strip()
    product     = (payload.get("product") or "MIS").strip().upper()    # MIS / CNC
    validity    = (payload.get("validity") or "DAY").strip().upper()
    trigger     = str(payload.get("trigger") or "0").strip()
    pin         = (payload.get("pin") or "").strip()

    ts = now_ist().strftime("%Y-%m-%d %H:%M:%S IST")
    log_entry = {
        "timestamp": ts, "symbol": symbol, "side": side, "qty": qty,
        "order_type": order_type, "price": price, "product": product,
        "validity": validity, "trigger": trigger,
        "status": "REJECTED", "kotak_order_id": None, "message": "",
    }

    # Validation
    expected_pin = os.getenv("ORDER_PIN", "").strip()
    if expected_pin and pin != expected_pin:
        log_entry["message"] = "Wrong PIN"
        append_order(log_entry)
        return jsonify({"ok": False, "error": "Wrong PIN"}), 401

    scrip = find_scrip(symbol)
    if not scrip:
        log_entry["message"] = "Unknown symbol"
        append_order(log_entry)
        return jsonify({"ok": False, "error": "Unknown symbol"}), 400
    if not scrip.get("tradeable"):
        log_entry["message"] = "Symbol not tradeable (index)"
        append_order(log_entry)
        return jsonify({"ok": False, "error": "Index symbols can't be traded as cash"}), 400
    if side not in ("B", "S"):
        log_entry["message"] = "Side must be B or S"
        append_order(log_entry)
        return jsonify({"ok": False, "error": "Side must be B or S"}), 400
    try:
        if int(qty) <= 0:
            raise ValueError("qty<=0")
    except (TypeError, ValueError):
        log_entry["message"] = "Qty must be a positive integer"
        append_order(log_entry)
        return jsonify({"ok": False, "error": "Qty must be a positive integer"}), 400
    if order_type == "L":
        try:
            if float(price) <= 0:
                raise ValueError("price<=0")
        except (TypeError, ValueError):
            log_entry["message"] = "Price required for LIMIT"
            append_order(log_entry)
            return jsonify({"ok": False, "error": "Price required for LIMIT order"}), 400

    # Login (must happen before margin fetch + safe wrapper)
    try:
        client = ensure_client()
    except Exception as e:
        log_entry["message"] = f"login: {e}"
        append_order(log_entry)
        return jsonify({"ok": False, "error": str(e)}), 500

    # Fetch available cash for the margin pre-check inside place_order_safe.
    # Best-effort: if Kotak limits() fails, we just skip the check (wrapper
    # treats available_cash=None as 'skip').
    available_cash = None
    try:
        ld, _ = safe_call(client.limits, segment="ALL", exchange="ALL", product="ALL")
        if isinstance(ld, dict):
            for k in ("Net", "net", "CashAvailable", "cashAvailable",
                      "AvailableCash", "availableCash", "DepositValue"):
                if k in ld:
                    try:
                        available_cash = float(ld[k]); break
                    except (TypeError, ValueError):
                        pass
    except Exception:
        pass

    # Single safe entry-point. Handles LIVE_MODE / kill switch / margin /
    # broker call / response parsing in one place. See backend/safety/orders.py.
    res = place_order_safe(
        client=client, scrip=scrip, side=side, qty=qty, price=price,
        order_type=order_type, product=product, validity=validity,
        trigger=trigger, tag="gann-ui",
        live_mode=LIVE_MODE, available_cash=available_cash,
        lot_size=1, source="manual_ticket",
    )

    # Persist a row to orders_log.json (for the /orderlog UI). Map wrapper
    # result -> the existing log_entry shape so the orderlog page stays the same.
    if res["result"] == RESULT_OK:
        log_entry["status"] = "PLACED"
        log_entry["kotak_order_id"] = res["order_id"]
        log_entry["message"] = res["message"]
        append_order(log_entry)
        return jsonify({"ok": True, "order_id": res["order_id"],
                        "message": res["message"], "raw": res["raw"]})

    if res["result"] == RESULT_PAPER:
        # LIVE_MODE is False — the manual ticket was clicked but we're in paper
        # mode. Surface a clear message so the operator isn't confused into
        # thinking a real order went through.
        log_entry["status"] = "PAPER"
        log_entry["message"] = res["message"]
        append_order(log_entry)
        return jsonify({"ok": False, "error": res["message"],
                        "paper_mode": True}), 400

    if res["result"] == RESULT_BLOCKED_HALTED:
        log_entry["status"] = "HALTED"
        log_entry["message"] = res["message"]
        append_order(log_entry)
        return jsonify({"ok": False, "error": res["message"],
                        "halted": True}), 423  # 423 Locked

    if res["result"] == RESULT_BLOCKED_MARGIN:
        log_entry["status"] = "INSUFFICIENT_FUNDS"
        log_entry["message"] = res["message"]
        append_order(log_entry)
        return jsonify({"ok": False, "error": res["message"]}), 402  # Payment Required

    # RESULT_KOTAK_ERROR or anything unrecognised
    log_entry["status"] = "REJECTED"
    log_entry["message"] = res["message"]
    append_order(log_entry)
    return jsonify({"ok": False, "error": res["message"],
                    "raw": res["raw"]}), 400


@app.route("/orderlog")
def orderlog_view():
    return render_template(
        "orderlog.html", tabs=TABS, active="orderlog",
        orders=read_orders(),
    )


@app.route("/orderlog.csv")
def orderlog_csv():
    rows = read_orders()
    cols = ["timestamp", "symbol", "side", "qty", "order_type", "price",
            "product", "validity", "trigger", "status", "kotak_order_id", "message"]
    lines = [",".join(cols)]
    for r in rows:
        vals = []
        for c in cols:
            v = str(r.get(c, ""))
            if "," in v or '"' in v or "\n" in v:
                v = '"' + v.replace('"', '""') + '"'
            vals.append(v)
        lines.append(",".join(vals))
    return Response("\n".join(lines), mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=orders_log.csv"})


@app.route("/audit")
def audit_view():
    """Audit events — paginated newest-first. Useful for forensics after
    a halt or unexpected behaviour. Query params: page=N, date=YYYY-MM-DD,
    event=PLACE_ORDER_OK (F.4 — narrows to a single event type).
    Server-side paginated so the response stays small even when audit.log
    has grown to thousands of lines."""
    page = request.args.get("page", 1)
    date = request.args.get("date") or ""
    event = request.args.get("event") or ""
    p = read_audit_page(page=page, page_size=50, date=date, event=event)
    # I.2 — surface the on-disk audit.log size so Ganesh can spot when the
    # file is approaching rotate-territory (it's never auto-deleted). Cheap
    # os.stat — no log content read.
    try:
        from backend.safety.audit import AUDIT_FILE as _AUDIT_FILE
        _bytes = os.path.getsize(_AUDIT_FILE)
        if _bytes >= 1024 * 1024:
            log_size_hint = f"{_bytes / (1024 * 1024):.1f} MB"
        elif _bytes >= 1024:
            log_size_hint = f"{_bytes / 1024:.1f} KB"
        else:
            log_size_hint = f"{_bytes} B"
    except OSError:
        log_size_hint = None
    return render_template(
        "audit.html",
        tabs=TABS, active="audit",
        events=p["items"],
        pagination=p,
        date_filter=date,
        event_filter=event,
        distinct_events=p.get("distinct_events", []),
        log_size_hint=log_size_hint,
    )


# ---------- KILL SWITCH ----------
# /STOP shows a confirmation page (GET) so a fat-finger swipe doesn't halt
# trading. /STOP/confirm (POST) is what actually engages the halt. Re-arming
# is intentionally NOT a web action — Ganesh must SSH in and remove
# data/HALTED.flag manually after investigating.
@app.route("/STOP", methods=["GET"])
def stop_view():
    return render_template("stop_confirm.html",
                           tabs=TABS, active=None,
                           live_mode=LIVE_MODE,
                           halted=is_halted(),
                           halt_info=halt_info())


@app.route("/STOP/confirm", methods=["POST"])
def stop_confirm():
    reason = (request.form.get("reason") or "manual_web").strip()
    halt(reason=reason)
    audit("KILL_SWITCH_HALT", source="web", reason=reason)
    return redirect(url_for("stop_view"))


# Make LIVE_MODE + halt state + UCC + today's combined P&L available to every
# template. The safety header partial uses these so the LIVE pill, clock, UCC
# label and "today P&L" tile show on every page (B.2).

# Bug 4 — cache today_pnl for 5 s. The context processor runs on every page
# render; with many tabs open this adds up even after the D.2 ledger memo
# made the underlying reads cheap. The filter+sum still walks both ledgers
# in Python — collapsing repeated calls during a 5 s window cuts CPU on the
# request thread without losing freshness (the tile already updates on the
# next page nav, and a closed-position P&L change is not real-time anyway).
_TODAY_PNL_CACHE = {"ts": 0.0, "value": None, "by_engine": None}
_TODAY_PNL_LOCK = threading.Lock()
_TODAY_PNL_TTL = 5.0


def _compute_today_pnl():
    """Walk both ledgers once, return (combined_pnl, {engine: pnl}).
    Used by both _today_pnl_cached and the per-engine drawdown check
    so a single ledger pass populates both views."""
    from backend.storage.paper_ledger import read_paper_ledger
    real_today  = _filter_trades_by_range(read_trade_ledger(),  "today", "")
    paper_today = _filter_trades_by_range(read_paper_ledger(),  "today", "")
    combined = 0.0
    by_engine = {"current": 0.0, "reverse": 0.0}
    for t in (real_today + paper_today):
        if t.get("status") != "CLOSED" or t.get("pnl_points") is None:
            continue
        try:
            rs = float(t["pnl_points"]) * int(t.get("qty", 1))
        except (TypeError, ValueError):
            continue
        combined += rs
        eng = t.get("engine") or "current"
        if eng in by_engine:
            by_engine[eng] += rs
        else:
            # Future-proof: any new engine name shows up as its own bucket.
            by_engine[eng] = by_engine.get(eng, 0.0) + rs
    return round(combined, 2), {k: round(v, 2) for k, v in by_engine.items()}


def _today_pnl_cached():
    now = time.time()
    with _TODAY_PNL_LOCK:
        if (now - _TODAY_PNL_CACHE["ts"]) < _TODAY_PNL_TTL:
            return _TODAY_PNL_CACHE["value"]
    try:
        value, by_engine = _compute_today_pnl()
    except Exception:
        value, by_engine = None, None
    with _TODAY_PNL_LOCK:
        _TODAY_PNL_CACHE["ts"] = now
        _TODAY_PNL_CACHE["value"] = value
        _TODAY_PNL_CACHE["by_engine"] = by_engine
    return value


def _today_pnl_by_engine_cached():
    """Per-engine P&L for today, dict keyed by engine name. Shares the
    same 5s cache as _today_pnl_cached — calling either populates both."""
    _today_pnl_cached()
    with _TODAY_PNL_LOCK:
        return dict(_TODAY_PNL_CACHE["by_engine"] or {})


def _parse_halt_info(text):
    """Parse a halt-flag file body (key=value lines) into a small dict.
    Returns None for None/empty input. Tolerates missing keys — used by
    the /config halt banner (P1.1) so reason + timestamp display
    cleanly instead of as a raw multiline blob."""
    if not text:
        return None
    out = {"halted_at": "", "reason": "", "engine": ""}
    for line in text.splitlines():
        if "=" in line:
            k, v = line.split("=", 1)
            k = k.strip()
            if k in out:
                out[k] = v.strip()
    return out


@app.context_processor
def _inject_safety_state():
    # B.2 — today's combined P&L (real ledger + paper book), summed in points
    # × qty so it matches the per-page Trade Log "Net P&L (₹)" totals.
    # P1.1 — halt_global / halt_engines feed the /config halt banner.
    # Cheap to compute (3 stat() calls + tiny file reads); fine in the
    # context processor since every page already pays the is_halted() stat.
    return {
        "live_mode": LIVE_MODE,
        "halted": is_halted(),
        "halt_global": _parse_halt_info(halt_info()),
        "halt_engines": {
            eng: _parse_halt_info(engine_halt_info(eng))
            for eng in ("current", "reverse")
        },
        "ucc": os.getenv("KOTAK_UCC", ""),
        "today_pnl": _today_pnl_cached(),
    }


@app.route("/refresh", methods=["POST", "GET"])
def refresh():
    _state["client"] = None
    _state["login_time"] = None
    _state["greeting"] = None
    _state["error"] = None
    # Trigger fresh login immediately (don't wait for next page view)
    try:
        ensure_client()
    except Exception:
        pass  # error already recorded in _state and history
    return redirect(url_for("holdings_view"))


# ============================================================
# Autonomous strategy ticker — runs option_auto_strategy_tick()
# every 3 seconds during market hours, regardless of whether any
# browser has the Options page open.
#
# Why: in Phase 4 (LIVE), we cannot depend on a human keeping a
# tab open. The bot must see every tick autonomously. Browser
# polling still calls the same tick function (idempotent — guarded
# by the strategy's per-path lock) but is no longer the only
# trigger. - matha
# ============================================================
TICKER_INTERVAL_SECONDS = 3
_ticker_state = {"started": False}


def _check_drawdown_halt():
    """H.2 — auto-engage the kill switch if today's combined P&L breaches
    the configured drawdown threshold. Idempotent: a no-op when already
    halted, when the threshold is unset, or when P&L is healthy.

    Runs once per ticker iteration so the same threshold isn't tripped
    repeatedly within a single tick. Audit + halt-flag write happen
    exactly once per breach because is_halted() flips True after the
    first call and short-circuits subsequent ticks.

    Phase 3 — also checks per-engine thresholds. Each logic engine
    (current/reverse) has its own max_daily_drawdown in config; if
    that engine's P&L breaches its threshold and the OTHER engine is
    still healthy, only that engine gets halted (halt_engine) rather
    than the whole bot. The global threshold is unchanged — it still
    halts everything when combined P&L is bad enough.
    """
    try:
        # Always populate the per-engine cache so we get both views
        # off a single ledger pass, even when global is already halted.
        pnl = _today_pnl_cached()
        by_engine = _today_pnl_by_engine_cached()

        # 1. Global combined-P&L threshold (unchanged from H.2).
        if not is_halted():
            threshold = config_loader.max_daily_drawdown()
            if threshold and pnl is not None and pnl <= -float(threshold):
                # Reason string is ASCII-only — halt() opens the flag file
                # without an explicit encoding, and on Windows that's cp1252
                # which would crash on the rupee glyph. "Rs." matches the
                # convention already used in safety/orders.py messages.
                reason = (f"auto-drawdown: today P&L Rs.{pnl:.2f} <= "
                          f"-Rs.{int(threshold)} threshold")
                halt(reason=reason)
                try:
                    audit("AUTO_HALT_DRAWDOWN", today_pnl=pnl,
                          threshold=int(threshold))
                except Exception:
                    pass
                print(f"[ticker] {reason} — kill switch engaged")

        # 2. Per-engine threshold — independent of global. If the engine
        # is already halted (via global flag OR its own per-engine flag)
        # is_engine_halted short-circuits; we don't re-engage on every
        # tick. Per-engine halt blocks NEW entries from that engine
        # only — exits + square-off + the other engine keep running.
        for eng in ("current", "reverse"):
            if not config_loader.engine_enabled(eng):
                continue
            if is_engine_halted(eng):
                continue
            eng_threshold = config_loader.engine_max_daily_drawdown(eng)
            if not eng_threshold:
                continue
            eng_pnl = by_engine.get(eng) if by_engine else None
            if eng_pnl is None:
                continue
            if eng_pnl <= -float(eng_threshold):
                reason = (f"auto-drawdown ({eng}): engine P&L "
                          f"Rs.{eng_pnl:.2f} <= -Rs.{int(eng_threshold)} "
                          f"threshold")
                halt_engine(eng, reason=reason)
                try:
                    audit("AUTO_HALT_ENGINE_DRAWDOWN", engine=eng,
                          engine_pnl=eng_pnl,
                          threshold=int(eng_threshold))
                except Exception:
                    pass
                print(f"[ticker] {reason} — {eng} engine halted")
    except Exception as e:
        # Best-effort guard. A drawdown-check failure must NOT kill the
        # ticker — strategy logic still needs to run for stoploss/exit.
        print(f"[ticker] drawdown check failed: {type(e).__name__}: {e}")


def _strategy_ticker_loop():
    """Daemon thread body. Sleep / check hours / tick / repeat. Errors are
    swallowed and logged so a single bad tick can't kill the loop."""
    print("[ticker] autonomous strategy ticker started "
          f"(every {TICKER_INTERVAL_SECONDS}s during 09:15-15:15 IST)")
    while True:
        try:
            now = now_ist()
            if _auto_in_hours(now):
                _check_drawdown_halt()  # H.2 — before each tick.
                try:
                    data, meta, _err = fetch_option_quotes()
                    if meta:
                        gann_quotes, _ = fetch_quotes()
                        try:
                            client_for_strategy = ensure_client()
                        except Exception:
                            client_for_strategy = None
                        # Phase 2c — run each enabled logic engine in
                        # turn. current + reverse are independent: own
                        # state slots, own config block, own per-engine
                        # ledger filtering. Each tick short-circuits
                        # internally if its master flag is off, so this
                        # loop is cheap when only one is enabled.
                        active_engines = [
                            e for e in ("current", "reverse")
                            if config_loader.engine_enabled(e)
                        ]
                        for _eng in active_engines:
                            try:
                                option_auto_strategy_tick(
                                    data, meta, gann_quotes,
                                    client=client_for_strategy,
                                    engine=_eng,
                                )
                            except Exception as e:
                                print(f"[ticker] real options tick "
                                      f"engine={_eng} failed: "
                                      f"{type(e).__name__}: {e}")
                            # Paper book — runs the same strategy logic
                            # against an independent ledger. Never sends
                            # real orders; not gated by the kill switch.
                            try:
                                paper_options_tick(
                                    data, meta, gann_quotes,
                                    engine=_eng,
                                )
                            except Exception as e:
                                print(f"[ticker] paper options tick "
                                      f"engine={_eng} failed: "
                                      f"{type(e).__name__}: {e}")
                        # Futures runs alongside options. Independent ledger
                        # rows (asset_type=future). Fetch futures quotes
                        # only if at least one engine (paper OR real)
                        # wants futures — otherwise it's a wasted REST call.
                        if (config_loader.real_futures_enabled()
                                or config_loader.paper_futures_enabled()):
                            try:
                                fut_data, _fut_err = fetch_future_quotes()
                                if fut_data:
                                    for _eng in active_engines:
                                        try:
                                            future_auto_strategy_tick(
                                                fut_data, gann_quotes,
                                                client=client_for_strategy,
                                                engine=_eng,
                                            )
                                        except Exception as e:
                                            print(f"[ticker] real futures "
                                                  f"tick engine={_eng} "
                                                  f"failed: "
                                                  f"{type(e).__name__}: {e}")
                                        try:
                                            paper_futures_tick(
                                                fut_data, gann_quotes,
                                                engine=_eng,
                                            )
                                        except Exception as e:
                                            print(f"[ticker] paper futures "
                                                  f"tick engine={_eng} "
                                                  f"failed: "
                                                  f"{type(e).__name__}: {e}")
                            except Exception as e:
                                print(f"[ticker] futures tick failed: "
                                      f"{type(e).__name__}: {e}")
                except Exception as e:
                    print(f"[ticker] tick failed: "
                          f"{type(e).__name__}: {e}")
        except Exception as e:
            # Outermost guard — should be unreachable but keeps the
            # thread alive even on truly unexpected failures.
            print(f"[ticker] loop guard caught: "
                  f"{type(e).__name__}: {e}")
        time.sleep(TICKER_INTERVAL_SECONDS)


def _start_strategy_ticker_once():
    """Idempotent — safe to call from multiple boot paths."""
    if _ticker_state["started"]:
        return
    _ticker_state["started"] = True
    t = threading.Thread(target=_strategy_ticker_loop,
                         daemon=True, name="strategy-ticker")
    t.start()


def _preload_option_universe():
    """Warm up the F&O universe cache in background so first /options visit
    is fast instead of waiting 90s for 3 sequential search_scrip calls."""
    _option_universe["preload_status"] = {}
    for idx_name in INDEX_OPTIONS_CONFIG:
        try:
            items, err = _fetch_index_fo_universe(idx_name)
            msg = f"{len(items)} contracts" + (f" (err: {err})" if err else "")
            _option_universe["preload_status"][idx_name] = msg
            print(f"[options] preloaded {idx_name}: {msg}")
        except Exception as e:
            msg = f"EXCEPTION: {type(e).__name__}: {e}"
            _option_universe["preload_status"][idx_name] = msg
            print(f"[options] preload {idx_name} failed: {msg}")


if __name__ == "__main__":
    print("=" * 60)
    print("Kotak Neo Dashboard starting...")
    mode = "LIVE (real money)" if LIVE_MODE else "PAPER (no real orders)"
    print(f"Trading mode: {mode}")
    if is_halted():
        print("!! KILL SWITCH IS ENGAGED — new live orders will be refused.")
    print("Open http://localhost:5000 in your browser")
    print("=" * 60)
    # Bug 6 — fail fast if another process is already bound to :5000.
    # Otherwise we'd start the strategy ticker + snapshot producer threads,
    # then crash inside app.run() leaving the user confused (and on Windows
    # leaking those daemon threads into the orphaned process). Probe the
    # port FIRST so the operator sees a clear "kill the old process" message
    # before any background work spins up.
    import socket as _socket
    _probe = _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM)
    try:
        _probe.settimeout(0.5)
        if _probe.connect_ex(("127.0.0.1", 5000)) == 0:
            print("!! Port 5000 is already in use.")
            print("   Another Kotak Neo Dashboard process is already running,")
            print("   or a previous run did not shut down cleanly.")
            print("   On Windows: taskkill /F /IM python.exe   (kills ALL python)")
            print("   Then start this app again. Aborting.")
            raise SystemExit(1)
    finally:
        _probe.close()
    audit("MODE_STARTUP", live_mode=LIVE_MODE, halted=is_halted())
    # Kick off the autonomous strategy ticker BEFORE app.run blocks. Daemon
    # thread, dies with the process on shutdown.
    _start_strategy_ticker_once()
    # Start the SnapshotStore producer — pre-builds /api/option-prices-v2
    # payload bytes every 2s so HTTP requests are O(1) reads.
    _snapshot.start()
    # threaded=True: don't block other requests while a slow search_scrip runs
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
